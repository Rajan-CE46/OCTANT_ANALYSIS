from openpyxl import load_workbook
import openpyxl
import math
import numpy as np
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import re


def process_data(df, mod=5000):
    # mod = 4000
    # df = pd.read_csv('octant_input.csv')
    df.insert(4, 'U_Avg', None)
    df.insert(5, 'V_Avg', None)
    df.insert(6, 'W_Avg', None)
    df.insert(7, "U'=U-U Avg", None)
    df.insert(8, "V'=V-V Avg", None)
    df.insert(9, "W'=W-W Avg", None)
    df.insert(10, 'Octant', None)

    emptyCol = ' '
    for i in range(41):
        df.insert(11, emptyCol, None)
        emptyCol = emptyCol + ' '

    df = df.fillna(' ')

    df.iat[0, 13] = 'Octant Count'
    df.iat[0, 23] = 'Ranking of Octant'
    df.iat[0, 34] = 'Transition Count'

    df.iat[2, 12] = f'Mod {mod}'
    df.iat[0, 4] = df['U'].mean()
    df.iat[0, 5] = df['V'].mean()
    df.iat[0, 6] = df['W'].mean()

    dfLen = len(df)

    def identifyOctant(u, v, w):
        if (u >= 0 and v >= 0 and w >= 0):
            return 1
        elif (u >= 0 and v >= 0 and w < 0):
            return -1
        elif (u < 0 and v >= 0 and w >= 0):
            return 2
        elif (u < 0 and v >= 0 and w < 0):
            return -2
        elif (u < 0 and v < 0 and w >= 0):
            return 3
        elif (u < 0 and v < 0 and w < 0):
            return -3
        elif (u >= 0 and v < 0 and w >= 0):
            return 4
        elif (u >= 0 and v < 0 and w < 0):
            return -4

    for i in range(dfLen):
        df.iat[i, 7] = df.iat[i, 1] - df.iat[0, 4]
        df.iat[i, 8] = df.iat[i, 2] - df.iat[0, 5]
        df.iat[i, 9] = df.iat[i, 3] - df.iat[0, 6]
        uEff = df.iat[i, 7]
        vEff = df.iat[i, 8]
        wEff = df.iat[i, 9]
        df.iat[i, 10] = identifyOctant(uEff, vEff, wEff)

    map_octantCount = {1: [], -1: [], 2: [], -
                       2: [], 3: [], -3: [], 4: [], -4: []}
    octantID = [1, -1, 2, -2, 3, -3, 4, -4]
    rangeDiv = math.ceil(dfLen/mod)

    map_TransitionCount = {}
    for id1 in octantID:
        for id2 in octantID:
            map_TransitionCount[(id1, id2)] = []

    for rangeFactor in range(0, rangeDiv):
        startingRow = (mod)*rangeFactor
        octCountInRange = {1: 0, -1: 0, 2: 0, -2: 0, 3: 0, -3: 0, 4: 0, -4: 0}
        map_TransitionCountInRange = {}
        for id1 in octantID:
            for id2 in octantID:
                map_TransitionCountInRange[(id1, id2)] = 0

        for row in range(startingRow, min(dfLen, startingRow+mod)):
            octCountInRange[df.iat[row, 10]] += 1
            if (row+1 < dfLen):
                map_TransitionCountInRange[(
                    df.iat[row, 10], df.iat[row+1, 10])] += 1
        for ID in octantID:
            map_octantCount[ID].append(octCountInRange[ID])
        for id1 in octantID:
            for id2 in octantID:
                map_TransitionCount[(id1, id2)].append(
                    map_TransitionCountInRange[(id1, id2)])

    # print(map_TransitionCount)

    low = mod-1
    for row in range(1, 1+rangeDiv+2):
        if (row == 1):
            df.iat[row, 13] = 'Octant ID'
        elif (row == 2):
            df.iat[row, 13] = 'Overall Count'
        elif (row == 3):
            df.iat[row, 13] = f'.0000-{mod-1}'
        else:
            high = low + mod
            low = low + 1
            df.iat[row, 13] = f'{low}-{min(high,dfLen-1)}'
            low = high

    for col in range(14, 22):
        df.iat[1, col] = octantID[col-14]
    for col in range(14, 22):
        for row in range(2, rangeDiv+2+1):
            if (row == 2):
                df.iat[row, col] = sum(map_octantCount[df.iat[1, col]])
            else:
                df.iat[row, col] = map_octantCount[df.iat[1, col]][row-3]

    columnIdMap = {}

    for col in range(23, 31):
        df.iat[1, col] = f'Rank of Octant {octantID[col-23]}'
        columnIdMap[octantID[col-23]] = col

    df.iat[1, 31] = 'Rank1 Octant ID'
    df.iat[1, 32] = 'Rank1 Octant Name'

    octantNameIdMapping = {
        1: "Internal outward interaction",
        -1: "External outward interaction",
        2: "External Ejection",
        -2: "Internal Ejection",
        3: "External inward interaction",
        -3: "Internal inward interaction",
        4: "Internal sweep",
        -4: "External sweep"
    }

    rank1IDCnt = {1: 0, -1: 0, 2: 0, -2: 0, 3: 0, -3: 0, 4: 0, -4: 0}
    for i in range(rangeDiv+1):
        countMap = {}
        countList = []
        for col in range(14, 22):
            countMap[df.iat[2+i, col]] = df.iat[1, col]
            countList.append(df.iat[2+i, col])
        countList.sort(reverse=True)
        rank = 1
        rank1IDCnt[countMap[countList[0]]] += 1
        df.iat[2+i, 31] = countMap[countList[0]]

        df.iat[2+i, 32] = octantNameIdMapping[countMap[countList[0]]]
        for cnt in countList:
            df.iat[2+i, columnIdMap[countMap[cnt]]] = rank
            rank += 1

    for row in range(2+rangeDiv+4, rangeDiv+6+9):
        if (row == rangeDiv+6):
            df.iat[row, 26] = 'Octant ID'
            df.iat[row, 27] = 'Octant Name'
            df.iat[row, 28] = 'Count of Rank 1 Mod Values'
        else:
            df.iat[row, 26] = octantID[row - (rangeDiv+4+2+1)]
            df.iat[row,
                   27] = octantNameIdMapping[octantID[row - (rangeDiv+4+2+1)]]
            df.iat[row, 28] = rank1IDCnt[octantID[row - (rangeDiv+4+2+1)]]

    df.iat[0, 34] = 'Transtion Details:'

    lower = 0
    higher = 0
    for section in range(rangeDiv+1):
        if (section == 0):
            df.iat[0, 35] = 'Overall Transtion Count'
            df.iat[2, 34] = 'From'
            df.iat[0, 36] = 'To'
            df.iat[1, 35] = 'Octant ID'
            for drow in range(2, 10):
                df.iat[drow, 35] = octantID[drow-2]
                fromID = octantID[drow-2]
                for ToID in range(36, 44):
                    df.iat[drow, ToID] = sum(
                        map_TransitionCount[(fromID, octantID[ToID-36])])
            for dcol in range(36, 44):
                df.iat[1, dcol] = octantID[dcol-36]
        else:
            higher = lower + mod - 1
            startOfCurrSection = section*13
            df.iat[startOfCurrSection, 35] = 'Mod Transtion Count Between: '
            df.iat[startOfCurrSection+1, 36] = 'To'
            df.iat[startOfCurrSection+1, 35] = f'{lower}-{min(dfLen-1,higher)}'
            lower = higher + 1
            df.iat[startOfCurrSection+2, 35] = 'Octant ID'
            df.iat[startOfCurrSection+3, 34] = 'From'

            for drow in range(startOfCurrSection+3, startOfCurrSection+11):
                df.iat[drow, 35] = octantID[drow-(startOfCurrSection+3)]
                fromID = octantID[drow-(startOfCurrSection+3)]
                for ToID in range(36, 44):
                    df.iat[drow, ToID] = map_TransitionCount[(
                        fromID, octantID[ToID-36])][section-1]
            for dcol in range(36, 44):
                df.iat[startOfCurrSection+2, dcol] = octantID[dcol-36]

    map_LongestSubsequence = {1: [0, []], -1: [0, []], 2: [0, []], -
                              2: [0, []], 3: [0, []], -3: [0, []], 4: [0, []], -4: [0, []]}
    for i in range(dfLen):
        currOct = df.iat[i, 10]
        currentMaxLen = map_LongestSubsequence[currOct][0]
        cnt = 1
        startTime = df.iat[i, 0]
        endTime = df.iat[i, 0]
        while (i+1 < dfLen and currOct == df.iat[i+1, 10]):
            cnt += 1
            endTime = df.iat[i, 0]
            i += 1
        i -= 1

        if (cnt > currentMaxLen):
            map_LongestSubsequence[currOct] = [cnt, [(startTime, endTime)]]
        elif (cnt == currentMaxLen):
            map_LongestSubsequence[currOct][1].append((startTime, endTime))

    df.iat[0, 45] = 'Longest Subsequence Details'

    for row in range(1, 10):
        if (row == 1):
            df.iat[row, 45] = 'Octant ID'
            df.iat[row, 46] = 'Longest Subsequence Length'
            df.iat[row, 47] = 'Count'
        else:
            df.iat[row, 45] = octantID[row-2]
            df.iat[row, 46] = map_LongestSubsequence[octantID[row-2]][0]
            df.iat[row, 47] = len(map_LongestSubsequence[octantID[row-2]][1])

    df.iat[1, 49] = 'Octant ID'
    df.iat[1, 50] = 'Longest Subsequence Length'
    df.iat[1, 51] = 'Count'

    row = 2
    for id in octantID:
        df.iat[row, 49] = id
        df.iat[row, 50] = map_LongestSubsequence[id][0]
        df.iat[row, 51] = len(map_LongestSubsequence[id][1])
        df.iat[row+1, 49] = 'Time'
        df.iat[row+1, 50] = 'From'
        df.iat[row+1, 51] = 'To'
        row += 2
        for times in range(len(map_LongestSubsequence[id][1])):
            df.iat[row, 50] = map_LongestSubsequence[id][1][times][0]
            df.iat[row, 51] = map_LongestSubsequence[id][1][times][1]
            row += 1

    # print(map_LongestSubsequence)
    return df

    # df.to_csv("octant_output.csv", index=False)


# def main():
#     df = pd.read_excel('5.3.xlsx')
#     result_df = process_data(df, 5000)

#     result_df.to_csv('output.csv', index=False)
#     df_csv = pd.read_csv('output.csv')


#     # Convert DataFrame to Excel
#     df_csv.to_excel('output.xlsx', index=False)

#     # Open the Excel file
#     wb = load_workbook('output.xlsx')
#     ws = wb.active

#     # Regular expression to match positive and negative numbers
#     pattern = re.compile(r'^-?\d*\.?\d+$')

#     # Iterate over each cell and convert text-formatted numbers to numeric values
#     for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#         for cell in row:
#             # Check if the cell contains a text-formatted number
#             if isinstance(cell.value, str) and pattern.match(cell.value):
#                 # Convert the text-formatted number to a float
#                 cell.value = float(cell.value)
#     for col in ws.columns:
#         column = col[0].column_letter  # Get the column letter
#         ws.column_dimensions[column].width = 12 # Set the column width to accommodate up to 6 characters

#     # Save the Excel file
#     wb.save('output.xlsx')
    

# if __name__ == "__main__":
#     main()
